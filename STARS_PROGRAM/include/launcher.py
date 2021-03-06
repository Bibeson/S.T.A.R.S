import serial
from threading import Event, Thread
import serial.tools.list_ports
import random
import sys
import time
import multiprocessing as mp
from openpyxl import load_workbook, Workbook
from guizero import App, Text, Box
from openpyxl.chart import (
    ScatterChart,
    Reference,
    Series,
)
# from gpiozero import LED


class data_object:
    def __init__(self, *args):
        self.args = args

class Stack:
    def __init__(self):
        self.items = []

    def isEmpty(self):
        return self.items == []

    def push(self, item):
        self.items.append(item)

    def pop(self):
        return self.items.pop()

    def peek(self):
        if not (self.isEmpty()):
            return self.items[len(self.items) - 1]
        else:
            return None

    def size(self):
        return len(self.items)

getMegaDataStack = Stack()
send_mega_stack = Stack()

def gpio_blinker(color, loop_count, Pi):
    if Pi:
        if loop_count % 2 == 0:
            color.on()
        else:
            color.off()

def loop_counter(loop_number):
    loop_number += 1
    if loop_number >= 10:
        loop_number = 1
    return loop_number
# _______PITCH AND YAW THREAD________ #


def findMEGA():
    ports = list(serial.tools.list_ports.comports())
    for p in ports:
        if "0042" in p[2]:
            return (p[0])
    return ('NULL')


def MegaData(MEGA, send_flag, send_data, send_mega_main, kill_event, close_event):
    # Variables:

    mega_data = data_object

    while not kill_event.is_set() and not close_event.isSet():
        getdata = False
        if MEGA.is_open:
            try:
                startMarker = MEGA.read().decode()
            except Exception as e:
                print('[MegaDataThread] : ERROR: ' + str(e))
            else:
                if startMarker == "<":
                    megaDataStr = MEGA.read(25)  # READ DATA FORMATTED AS ('< 1 2 3 4 5 >')
                    startTime = time.time()
                    megaDataTemp = list(megaDataStr.decode())
                    megaDataTemp.insert(0, startMarker)
                    megaData = megaDataTemp[:megaDataTemp.index(">") + 1]
                    tempData = "".join(megaData)
                    getdata = True
                    # IF SEND FLAG IS SET, DATA IS SENT AFTER:
                    if send_flag.is_set():
                        try:
                            data = send_data.pop()
#                            print("[MegaData]", data)
                            # MEGA.reset_output_buffer()
                            MEGA.write(data.encode())
                            print("[MegaDataThread] : SENT MEGA DATA:  " + data)
                        except serial.SerialException as e:
                            print("[MegaDataThread] : ** MEGA SEND FAILED **   " + str(e))
                            while (time.time() - startTime >= 0.5):
                                try:
                                    #MEGA.reset_output_buffer()
                                    MEGA.write(data.encode())
                                    print("[MegaDataThread] : Launch motors starting...")
                                except serial.SerialException as e:
                                    print("[MegaDataThread] : ** MEGA SEND FAILED **   " + str(e))
                                    if time.time() - startTime >= 0.5:
                                        pass
                                else:
                                    send_flag.clear()
                        except:
                            print("[MegaDataThread] : Issue sending data")
                            pass
                        else:
                            send_flag.clear()

            finally:
                if getdata == True:
                    try:
                        tempData = tempData.strip("<")
                        tempData = tempData.strip(">")
                        tempData = tempData.split(",")
                        print("[MegaDataThread] : Tempdata=  " + str(tempData))
                        
                        lidar_2_Distance = int(tempData[0])  # lidarDistance = int(cm)
                        temperature = int(tempData[1])  # temperature = int()
                        voiceCommand = int(tempData[2])  # voice commands = int(from 1 to 5)
                        targetTiming = int(tempData[3])  # targetTiming = float(0.0)
                        targetBallSpeed = float(tempData[4])  # targetBallSpeed

                        mega_data.lidar_2_Distance = lidar_2_Distance
                        mega_data.temperature = temperature
                        mega_data.voiceCommand = voiceCommand
                        mega_data.targetTiming = targetTiming
                        mega_data.targetBallSpeed = targetBallSpeed
                        
                        if targetTiming or targetBallSpeed != 0.0:
                            print("targetTiming: ",targetTiming," ||  targetBallSpeed: ",targetBallSpeed)

                        # Send to Launcher Parent (STACK)
                        getMegaDataStack.push(mega_data)
                        # Send to Main_File Process (QUEUE)
                        send_mega_main.put(tempData)
                    except Exception as e:
                        print("[MegaDataThread] : Parsing Data:  " + str(e))
        else:
            print("[MegaDataThread] : ** Mega didnt respond  **")
            time.sleep(0.01)
            continue




class Launcher(mp.Process):
    '''# _____________________________________LAUNCHER THREAD__________________________________#
Connects and communicates with the Arduino Mega: Launcher Motors, Ball Feeder, Wifi, Accelerometer? (Rangefinder?)
Launcher(Thread): ---> Arduino MEGA provide Angle values to both motors, request temperature's from the Uno, and distance measurements
RECEIVE from MEGA:
- Lidar_2_Distance
- temperature
- voiceCommand
- targetTiming

SEND to MEGA:
- MotorSpeed1
- MotorSpeed2
- targetChoice
- difficulty
- estimated_tof
- targetBallSpeed

Stack IO
RECEIVE from STACKS:
- guiStack
- getfutureDist
- stereo_pitch_launch
- sendfinalDistStack


SEND to STACKS:
- sendMegaDataStack
- sendLidar2Stack
- sendfinalDistStack
- sendTemperatureStack'''

    def __init__(self, gui_data, send_mega_data, get_final_dist, get_future_dist, pi_no_pi, led_color, kill_event, pause_event, py_reset, launch_event,voice_cont, drill_results, player_name):
        super(Launcher, self).__init__()
        self.send_mega_data = send_mega_data
        self.gui_data = gui_data
        self.get_final_dist = get_final_dist
        self.get_future_dist = get_future_dist
        self.send_flag = Event()
        self.working_on_the_Pi = pi_no_pi
        self.kill_event = kill_event
        self.pause_event = pause_event
        self.py_reset = py_reset
        self.launch_event = launch_event
        self.voice_control = voice_cont
        self.drill_result = drill_results
        self.player = player_name
        if self.working_on_the_Pi:
            self.color = led_color

        self.drillSpeed = 1
        self.difficulty = 1
        self.drillType = ""
        self.fasterVC = 1
        self.slowerVC = 2
        self.pauseVC = 3
        self.beginVC = 4
        self.changeDrillVC = 5
        self.changeManVC = 6
        self.drillCount = 1
        #self.MEGA = None
        self.OLD_FUT_FINAL_DIST = None
        self.FUT_FINAL_DIST = None
        self.LaunchTime = None
        self.launch_loop_count = 1
        self.stereoData = None
        self.stereo_Distance = 0.0
        self.rationaleDistMeasures = 0
        self.distanceTotal = 0
        self.used_distance = 0.0
        self.drill_data = data_object
        self.targetTimes = [0.0, 0.0, 0.0, 0.0, 0.0, 0.0]
        self.ballSpeeds = [0.0, 0.0, 0.0, 0.0, 0.0, 0.0]
        self.initiation = False

        # LAUNCHER OPTIONS:
        self.WAIT_TIME = 0
        self.DYNAMIC_WAIT_TIME = 20  # seconds
        self.STATIC_WAIT_TIME = 15  # seconds
        self.ballfeed = "0"
        # MEGA_DATA Stack values:
        self.voiceCommand = -1  # < set to 1 for testing the launcher (-1 otherwise)
        self.targetTiming = 0.0
        self.targetBallSpeed = 0.0
        self.startTime = 0.0
        self.first_drill = True
        # ____________________
        self.old_target_time = 0
        self.old_ballspeed = 0.0
        
    def drill_wait_time(self):
        if self.LaunchTime is not None:
            print("Entering Drill Waittime")
            while (time.time() - self.LaunchTime) < self.WAIT_TIME and not self.kill_event.is_set():
                if self.pause_event.is_set():
                    print("[Launcher] : Paused Drill")
                    while self.pause_event.is_set():
                        time.sleep(1)
                    
                self.first_drill = False
                #print("[Launcher] : Holding Loop for:  " + str(self.WAIT_TIME) + "  seconds  Waiting for results from target WIFI")
                time.sleep(.1)
                
                if self.kill_event.is_set():
                    self.close_drill()

                    sys.exit()
        elif self.drillCount ==1:
            print("[Launcher] : First Ball")
            self.first_drill = True

    def wait_for_voice(self):
        while not self.voiceCommand == self.beginVC and not self.kill_event.is_set():
            try:
                self.MEGA_DATA = getMegaDataStack.peek()
                self.voiceCommand = self.MEGA_DATA.voiceCommand
                self.temperature = self.MEGA_DATA.temperature
                if self.temperature != 0.0:
                    print("VC = " + str(self.voiceCommand) + "  TEMP = " + str(self.temperature))
                    self.drill_data.temp = self.temperature
                else:
                    print("VC = ", str(self.voiceCommand))
                time.sleep(0.5)
            except:
                print("[Launcher] : Waiting for Voice Command")
                time.sleep(1)
                continue

    def wait_for_target(self):
        target_data = False
        start = time.time()
        while not target_data and not self.kill_event.is_set():
            try:
                self.MEGA_DATA = getMegaDataStack.peek()
                self.voiceCommand = self.MEGA_DATA.voiceCommand
                self.temperature = self.MEGA_DATA.temperature
                self.targetTiming = self.MEGA_DATA.targetTiming
                self.targetBallSpeed = self.MEGA_DATA.targetBallSpeed
                if self.temperature != 0.0:
#                    print("VC = " + str(self.voiceCommand) + "  TEMP = " + str(self.temperature))
                    self.drill_data.temp = self.temperature
#                else:
#                    print("VC = ", str(self.voiceCommand))
                time.sleep(0.5)
            except:
                print("[Launcher] : Waiting for Voice Command")
                time.sleep(1)
                continue
            try:
#                print("In the try")
                i = self.drillCount -2
                if self.targetTiming != 0 and self.targetTiming != self.old_target_time:
                    if i != 0:
                        print("[Launcher] : targetTiming = " + str(self.targetTiming), " ||", i)
                    # self.drill_data.targetTime = self.targetTiming
                    # for i in range[4]:
                    self.targetTimes[i] = self.targetTiming
                    self.old_target_time = self.targetTiming
#                    self.targetTiming = 0  # < reset
                    target_data = True

                if self.targetBallSpeed != 0.0 and self.targetBallSpeed != self.old_ballspeed:
                    if i != 0:
                        print("[Launcher] : targetBallSpeed = " + str(self.targetBallSpeed)," ||", i)
                    # self.drill_data.ballSpeed = self.targetBallSpeed
                    # for i in range[4]:
                    self.ballSpeeds[i] = self.targetBallSpeed
                    self.old_ballspeed = self.targetBallSpeed
#                    self.targetBallSpeed = 0.0  # < reset
                    target_data = True

                if self.voiceCommand > 0:
                    if self.voiceCommand == self.pauseVC:
                        print("stop VC received")
                        self.kill_event.set()
                    elif self.voiceCommand == self.fasterVC:
                        print("faster VC received")
                        self.drillSpeed += 1
                    elif self.voiceCommand == self.slowerVC:
                        print("slower VC received")
                        self.drillSpeed -= 1
                    elif self.voiceCommand == self.changeDrillVC:
                        print("pause VC received")
                        self.pause_event.set()

                if self.pause_event.is_set():
                    print("[Launcher] : Paused Drill")
                    while self.pause_event.is_set():
                        time.sleep(1)

                end = time.time() - start
                if end > 10:
                    target_data = True
            except Exception as e:
                print("[Launcher]: ->",e)



    def get_mega_data(self):
        # _______ RECEIVE MEGA DATA (Wait): lidar_2_Distance, voiceCommand, targetTiming, targetBallSpeed _______#
        self.MEGA_DATA = getMegaDataStack.peek()
        self.lidar_2_Distance = self.MEGA_DATA.lidar_2_Distance  # lidarDistance = int(cm)
        self.temperature = self.MEGA_DATA.temperature  # temperature = int()
        self.voiceCommand = self.MEGA_DATA.voiceCommand  # voice commands = int(from 1 to 5)
        self.targetTiming = self.MEGA_DATA.targetTiming  # targetTiming = float(0.0)
        self.targetBallSpeed = self.MEGA_DATA.targetBallSpeed  # targetBallSpeed
        
        self.lidar_2_Distance = self.lidar_2_Distance / 100  # << convert LIDAR 2 from cm to meters
        #self.send_mega_data.put(self.MEGA_DATA)
        #print("[Launcher] : ","<" + str(self.lidar_2_Distance) + "," + str(self.temperature) + "," + str(self.voiceCommand) + "," + str(self.targetTiming) + "," + str(self.targetBallSpeed) + ">")

#        if self.targetTiming or self.targetBallSpeed != 0.0:
#            print("targetTiming: ", self.targetTiming, " ||  targetBallSpeed: ", self.targetBallSpeed)
        

    def send_launch_data(self):
        # ____POLYNOMIAL FIT FROM THEORETICAL VALUES___ #
        # ______________________ NEW SECTION FOR EACH PITCH ANGLE: ______________________ #
        if self.used_distance <= 7.5:
            RPM = -1.13635244 * self.used_distance ** 2.0 + 97.7378699 * self.used_distance + 646.034298  # <-- Polynomial fit
            RPS = RPM / 60
            PERIOD = (1 / RPS)*(1000000)
            print("[Launcher]: RPM: ", int(RPM), "  Period: ", int(PERIOD))

        elif 7.5 < self.used_distance <= 12.5:
            RPM = -1.13635244 * self.used_distance ** 2.0 + 97.7378699 * self.used_distance + 646.034298  # <-- Polynomial fit
            RPS = RPM / 60
            PERIOD = (1 / RPS) * (1000000) + 11500
            print("[Launcher]: RPM: ", int(RPM), "  Period: ", int(PERIOD))

        elif 12.5 < self.used_distance <=17.5:
            RPM = -1.13635244 * self.used_distance ** 2.0 + 97.7378699 * self.used_distance + 646.034298  # <-- Polynomial fit
            RPS = RPM / 60
            PERIOD = (1 / RPS) * (1000000) + 15000
            print("[Launcher]: RPM: ", int(RPM), "  Period: ", int(PERIOD))

        elif 17.5 < self.used_distance <= 22.5:
            RPM = -1.13635244 * self.used_distance ** 2.0 + 97.7378699 * self.used_distance + 646.034298  # <-- Polynomial fit
            RPS = RPM / 60
            PERIOD = (1 / RPS) * (1000000) + 10000
            print("[Launcher]: RPM: ", int(RPM), "  Period: ", int(PERIOD))

        elif 22.5 < self.used_distance <= 35.0:
            RPM = -1.13635244 * self.used_distance ** 2.0 + 97.7378699 * self.used_distance + 646.034298  # <-- Polynomial fit
            RPS = RPM / 60
            PERIOD = (1 / RPS) * (1000000) + 2500
            print("[Launcher]: RPM: ", int(RPM), "  Period: ", int(PERIOD))

        else: # << catches and strangeness that may happen to distance reading
            RPM = -1.13635244 * self.used_distance ** 2.0 + 97.7378699 * self.used_distance + 646.034298  # <-- Polynomial fit
            RPS = RPM / 60
            PERIOD = (1 / RPS) * (1000000)
            print("[Launcher]: RPM: ", int(RPM), "  Period: ", int(PERIOD))

        motor_period = str(int(PERIOD))
        targetChoice = int(random.choice([1,2]))
        estimated_tof = (0.120617 * self.used_distance)*1000  # + difficulty_time
        estimated_tof = round(estimated_tof,2)

        if self.first_drill and not self.initiation:
            time.sleep(5)
            print("Sending Target Initiation:")
            
            if self.drillType == "Dynamic":
                self.feed = "0"
                self.diff = "-1"
                drill_type = "1"
            else:
                self.feed = "0"
                self.diff = "-1"
                drill_type = "0"

            self.send_data = '<'+motor_period+','+motor_period+','+str(targetChoice)+','+str(self.diff)+','+self.feed+','+str(estimated_tof)+','+drill_type+'>'
#            print(self.send_data)
            send_mega_stack.push(self.send_data)
            # print("[Launcher] : BALL NUMBER :  ", self.drillCount)
            self.send_flag.set()
            while self.send_flag.is_set():
                time.sleep(0.1)
            self.initiation = True

            time.sleep(2)
        else:
            if self.drillType == "Dynamic":
                drill_type = "1"
            else:
                drill_type = "0"

            self.send_data = '<' + motor_period + ',' + motor_period + ',' + str(targetChoice) + ',' + str(
                self.difficulty) + ',' + self.ballfeed + ',' + str(estimated_tof) + ',' + drill_type + '>'

            # ____________________ Write data to MEGA ____________________
            send_mega_stack.push(self.send_data)
            print("[Launcher] : BALL NUMBER :  ", self.drillCount)
            ## NEW
            self.launch_event.set()
            ##_____
            self.send_flag.set()
            while self.send_flag.is_set():
                time.sleep(0.1)

            if self.ballfeed =="0":
                    time.sleep(0.2)
            if self.ballfeed =="1":
                print("[Launcher] : LAUNCHING BALL")
                self.drillCount += 1  # << ON SUCCESSFUL LAUNCH
                self.LaunchTime = time.time()

    def launcher_startup(self):
        startData = False
        while not startData and not self.kill_event.is_set() and not self.pause_event.is_set():
            self.drillSpeed = self.gui_data.speed
            self.difficulty = self.gui_data.difficulty
            self.drillType = self.gui_data.drilltype

            if self.pause_event.is_set():
                print("[Launcher] : Paused Drill")
                while self.pause_event.is_set():
                    time.sleep(1)

            try:
                mega_port = findMEGA()
                MEGA = serial.Serial(mega_port, 115200, timeout=1)
                MEGA.baudrate = 115200

                self.close_event = Event()
                # START COMMUNICATIONS THREAD
                self.mega_data_thread = Thread(target=MegaData,args=[MEGA, self.send_flag, send_mega_stack, self.send_mega_data, self.kill_event, self.close_event])
                self.mega_data_thread.start()

            except serial.SerialException as err:
                print("[Launcher] : MEGA thread failed to start" + str(err))
                time.sleep(2)
                continue
            except Exception as nodata:
                print("[Launcher] : ** Mega data thread failed to start:  " + str(nodata) + "**")
                time.sleep(2)
                # *****

                    # Get TEMPERATURE and WAIT for voiceCommand
            else:
                print("[Launcher] : Connected to MEGA")
                # Get drillType data from GUI__ #

                # print("LAUNCHER: Speed:  " + str(drillSpeed) + "  Diff:  " + str(difficulty) + "  Type:  " + str(drillType))

                # _____________________WAIT FOR VOICE COMMAND and TEMP TO BEGIN DRILL _____________________
                if self.voice_control.is_set():
                    self.wait_for_voice()
                else:
                    time.sleep(3)

                # ****SET TEMPERATURE CORRECTION FACTOR*****
                self.tempCorrect = 1  # self.temperature / 25
                # *****************************************
                startData = True

    def launcher_common(self):
        self.startTime = time.time()
        launch_data = False

        gpio_blinker(self.color, self.launch_loop_count, self.working_on_the_Pi)

        self.launch_loop_count = loop_counter(self.launch_loop_count)

        
        if not self.first_launch:
            self.wait_for_target()
        else:
            self.first_launch = False
            self.ballfeed="0"
            self.send_launch_data()
        
        self.drill_wait_time()

        while not launch_data and not self.kill_event.is_set():
            try:
                FINAL_DIST = self.get_final_dist.get(timeout=2)  # <<<<<
#                print("[Launcher] :  Final Dist : " + str(FINAL_DIST))
            except:
                print("[Launcher] : Waiting for final distance")
                time.sleep(2)
                continue
            else:
                self.used_distance = FINAL_DIST
                launch_data = True


        if self.pause_event.is_set():
            print("[Launcher] : Paused Drill")
            while self.pause_event.is_set():
                time.sleep(1)

    def dynamic_drill(self):
        # self.wait_for_voice()
        self.WAIT_TIME = self.DYNAMIC_WAIT_TIME
        print("[Launcher] : Beginning Drill ________________________")
        self.first_launch = True

        while self.drillCount <= 5 and not self.kill_event.is_set():
            if self.pause_event.is_set():
                print("[Launcher] : Paused Drill")
                while self.pause_event.is_set():
                    time.sleep(1)

            try:
                self.get_mega_data()
            except serial.SerialException as err:
                print("[Launcher] : MEGA not detected" + err)
                print("[Launcher] : Closing Launcher")
                self.close_drill()
            except Exception as e:  # (serial.SerialException)
                print("[Launcher] : Error in MegaData" + str(e))
                time.sleep(1)
                continue
            else:
                self.launcher_common()


                try:
                    self.FUT_FINAL_DIST = self.get_future_dist.get(timeout = 5)  # <<<<< GET PREDICTED LOCATION
                    print("[Launcher]: FUT_FINAL_DIST:  ",self.FUT_FINAL_DIST)
                except:
                    print("[Launcher] : Failed to get FUT FINAL DIST")
                    self.FUT_FINAL_DIST = None
                    pass
                finally:
                    if not self.FUT_FINAL_DIST:  # <- no prediction is done in this thread so it will send AS IS
                        print("[Launcher] : Sending launch data as is")
                        self.ballfeed = "0"
                        self.send_launch_data()

                        motor_start_time = time.time()

                        self.ballfeed = "1"

                        while time.time() - motor_start_time < 3:
                            time.sleep(0.2)

                        self.send_launch_data()
                    else:
                        self.used_distance = self.FUT_FINAL_DIST
                        self.ballfeed = "0"
                        self.send_launch_data()

                        motor_start_time = time.time()

                        self.ballfeed = "1"

                        while time.time() - motor_start_time < 3:
                            time.sleep(0.2)

                        self.send_launch_data()

        if self.drillCount >= 5:
            self.wait_for_target()
        self.close_drill()

    def static_drill(self):
        self.WAIT_TIME = self.STATIC_WAIT_TIME
        print("[Launcher] : Beginning Drill")
        self.first_launch = True

        while self.drillCount <= 5 and not self.kill_event.is_set():
            self.launcher_common()

            try:
                self.get_mega_data()
            except serial.SerialException as err:
                print("[Launcher] : MEGA not detected" + err)
                print("[Launcher] : Closing Launcher")
                self.close_drill()
            except Exception as e:  # (serial.SerialException)
                print("[Launcher] : Error in MegaData" + str(e))
                time.sleep(1)
                continue
            else:
                self.ballfeed = "1"

                self.send_launch_data()

        print("[Launcher] : Drill Ending  ")
        if self.drillCount >= 5:
            self.wait_for_target()
        self.close_drill()

    def manual_drill(self):
        self.WAIT_TIME = 0 # <<< WAIT TIME IS HANDLED DIFFERENTLY HERE
        print("[Launcher] : Beginning Drill")
        self.first_launch = True

        while self.drillCount <= 5 and not self.kill_event.is_set():
            # WAIT FOR VC FOR EVERY LAUNCh
            if not self.first_launch:
                self.wait_for_voice()

            self.launcher_common()

            try:
                self.get_mega_data()
            except:
                continue
            else:
                self.ballfeed = "1"
                self.send_launch_data()

        print("[Launcher] : Drill Ending  ")
        if self.drillCount >= 5:
            self.wait_for_target()
        self.close_drill()


    def output_to_excel(self):
        filename = "/home/pi/Desktop/demo_wb.xlsx"

        try:
            wb = load_workbook(filename=filename)
            print("[Launcher]: file exists")
        except Exception as e:
            print("[Launcher]: file not found, making new book ")
            wb = Workbook()
            summary_sheet = wb.get_sheet_by_name('Sheet')
            summary_sheet.title = 'Summary'

        try:
            players_sheet = wb[self.player]
            print("[Launcher]: sheet exists, appending data")

        except KeyError as k:
            print("[Launcher]: creating new sheet for: ", self.player)
            players_sheet = wb.create_sheet(title=self.player)

        data_printed = False
        col = 1
        row = 1
        drill = 1
        while not data_printed:
            if players_sheet.cell(column=col, row=row + 1).value is None:
                players_sheet.cell(column=col, row=row + 1).value = "Drill" + str(drill)
                if row == 1:
                    # players_sheet.cell(column=col, row=row).value = "."
                    players_sheet.cell(column=col + 1, row=row).value = "Ball #"
                    players_sheet.cell(column=col + 2, row=row).value = "Time for Pass"
                    players_sheet.cell(column=col + 3, row=row).value = "Ball Speed"
                for i in range(5):
                    players_sheet.cell(column=col + 1, row=row + i + 1).value = row + i
                    players_sheet.cell(column=col + 2, row=row + i + 1).value = (self.targetTimes[i + 1] / 1000)
                    players_sheet.cell(column=col + 3, row=row + i + 1).value = self.ballSpeeds[i + 1]

                data_printed = True

            else:
                row += 5
                drill += 1

        wb.save(filename)
        # PLOT THE DATA:
        ranges = [0] * (len(wb.worksheets) - 1)
        speed_ranges = [0] * (len(wb.worksheets) - 1)
        series = [0]*20
        speed_series = [0] * (len(wb.worksheets) - 1)
        
        chart1 = ScatterChart()
        chart2 = ScatterChart()
        chart3 = ScatterChart()
        chart4 = ScatterChart()
        chart5 = ScatterChart()
        charts = [chart1,chart2,chart3,chart4,chart5]
        
        speed_chart1 = ScatterChart()
        speed_chart2 = ScatterChart()
        speed_chart3 = ScatterChart()
        speed_chart4 = ScatterChart()
        speed_chart5 = ScatterChart()
        speed_charts = [speed_chart1,speed_chart2,speed_chart3,speed_chart4,speed_chart5]
        
        for i in range((len(wb.worksheets) - 1)):
#            print(str(wb.worksheets[i+1])[11:20])
            gathered = False
            coll = 1
            roww = 1
            speed_charts[i].title = str(wb.worksheets[i+1])[12:19] + "'s  Results"
            charts[i].title = str(wb.worksheets[i+1])[12:19] + "'s  Results"

            while not gathered:
                print("Gathering..")
                if wb.worksheets[i + 1].cell(column=coll, row=roww + 1).value is None:
                    gathered = True
                else:
                    roww += 5
                    gathered = False

            ranges[i] = Reference(wb.worksheets[i + 1], min_col=3, min_row=1, max_col=3,max_row=roww)
            x_vals = Reference(wb.worksheets[i + 1], min_col=2, min_row=1, max_row=roww)
            series[i] = Series(ranges[i], xvalues=x_vals, title_from_data=True)
            charts[i].append(series[0])
            
            charts[i].x_axis.title = "Test Number"
            charts[i].x_axis.scaling.min = 0
            charts[i].x_axis.scaling.max = roww
            charts[i].y_axis.title = "Reaction Times (seconds)"
            charts[i].legend = None

            s1 = charts[i].series[0]
            s1.marker.symbol = "triangle"
            s1.marker.graphicalProperties.solidFill = "FF0000"  # Marker filling
            s1.marker.graphicalProperties.line.solidFill = "FF0000"  # Marker outline
            s1.graphicalProperties.line.noFill = True

            speed_ranges[i] = Reference(wb.worksheets[i + 1], min_col=4, min_row=1, max_col=4, max_row=roww)
            x_value = Reference(wb.worksheets[i + 1], min_col=2, min_row=1, max_row=roww)
            series[i+10] = Series(speed_ranges[i], xvalues=x_value, title_from_data=True)
            speed_charts[i].append(series[i+10])
            
            speed_charts[i].x_axis.title = "Test Number"
            speed_charts[i].x_axis.scaling.min = 0
            speed_charts[i].x_axis.scaling.max = roww
            speed_charts[i].y_axis.title = "Ball Speed (m/s)"
            speed_charts[i].legend = None

            s2 = speed_charts[i].series[0]
            s2.marker.symbol = "triangle"
            s2.marker.graphicalProperties.solidFill = "006868"  # Marker filling
            s2.marker.graphicalProperties.line.solidFill = "006868"  # Marker outline
            s2.graphicalProperties.line.noFill = True
            
            charts[i].style = 13
            wb.worksheets[0].add_chart(charts[i], "A" + str(i + 1 + 15 * i))
            speed_charts[i].style = 13
            wb.worksheets[0].add_chart(speed_charts[i], "J" + str(i + 1 + 15 * i))

        wb.save(filename)

    def display_results(self):
        results = [0,0]
        results[0] = self.targetTimes
        results[1] = self.ballSpeeds
#        print(results)
#        results.append(self.player)
        self.drill_result.put(results)
        # ball = [0,0,0,0,0,0]
        # app = App(title="Results", layout="auto", height=320, width=480, bg="#424242", visible=True)
        # title = Text(app, str(self.player) + "'s Results for " + str(self.drillType), size=20, font="Calibri Bold", color="white")
        # subt = "Pass Difficulty: " + str(self.difficulty) + "  |  Ball Speed: " + str(self.drillSpeed)
        # subtitle = Text(app, subt, size=14, font="Calibri Bold", color="white")
        # spacer = Box(app, width=20, height=20)
        #
        # for i in range(5):
        #
        #     ball[i + 1] = "Pass # " + str(i + 1) + ":   " + str((self.targetTimes[i + 1]) / 1000) + " sec  |  " + str(
        #         self.ballSpeeds[i + 1]) + " m/s"
        #     # if len(ball[i + 1]) != 25:
        #     # ball[i+1] += "0"
        #     Text(app, ball[i + 1], size=12, font="Calibri Bold", color="white")
        #
        # app.display()

    def close_drill(self):
        print("[Launcher] : Drill Ending  ")
        if self.kill_event.is_set():
            print("[Launcher] :  Closing process")
            self.send_data = '<-1,-1,0,6,0,0.0>'
            send_mega_stack.push(self.send_data)
            self.send_flag.set()
            while self.send_flag.is_set() or self.py_reset.is_set:
                time.sleep(0.1)
            sys.exit()

        elif self.drillCount >= 5:
            last_data = False
            start = time.time()
            i = self.drillCount -2
            
#            print(i)
#            print(self.targetTimes)
#            print(self.ballSpeeds)
            while not last_data:
                try:
                    i = 5
                    self.MEGA_DATA = getMegaDataStack.peek()
                    self.voiceCommand = self.MEGA_DATA.voiceCommand
                    self.temperature = self.MEGA_DATA.temperature
                    self.targetTiming = self.MEGA_DATA.targetTiming
                    self.targetBallSpeed = self.MEGA_DATA.targetBallSpeed
                    
                    if self.temperature != 0.0:
#                        print("VC = " + str(self.voiceCommand) + "  TEMP = " + str(self.temperature))
                        self.drill_data.temp = self.temperature
#                    else:
#                        print("VC = ", str(self.voiceCommand))
                        
                    time.sleep(0.5)
                except:
                    print("[Launcher] : Waiting for Voice Command")
                    time.sleep(1)
                    continue
                else:
                    try:
                        print("Waiting for ball 5 target data")
                        if self.targetTiming !=  0 and self.targetTiming != self.old_target_time:
                            # self.drill_data.targetTime = self.targetTiming
                            # for i in range[4]:
                            self.targetTimes[i] = self.targetTiming
                            self.old_target_time = self.targetTiming
#                            self.targetTiming = 0  # < reset
                            if self.targetTimes[4] != self.targetTimes[5]:
                                print("[Launcher] : targetTiming = " , str(self.targetTiming)," |last| ", i)
                                last_data = True

                        if self.targetBallSpeed != 0.0 and self.targetBallSpeed != self.old_ballspeed:
                            # self.drill_data.ballSpeed = self.targetBallSpeed
                            # for i in range[4]:
                            self.ballSpeeds[i] = self.targetBallSpeed
#                            self.targetBallSpeed = 0.0  # < reset
                            self.old_ballspeed = self.targetBallSpeed
                            if self.targetTimes[4] != self.targetTimes[5]:
                                print("[Launcher] : targetBallSpeed = " , str(self.targetBallSpeed)," |last| ", i)
                                last_data = True
                            
                        end = time.time() - start

                        if end > 20:
                            last_data = True
                    except Exception as e:
                        print("[Launcher]: close_drill error", e)

#            while self.send_flag.is_set() or self.py_reset.is_set:
#                if self.send_flag.is_set():
#                    print('Send flag is set')
#                    time.sleep(15)
#                if self.py_reset.is_set():
#                    print('Reset flag is set')
#                    time.sleep(1)

            time.sleep(5)
            self.send_data = '<-1,-1,0,6,0,0.0>'
            send_mega_stack.push(self.send_data)
            self.send_flag.set()
            while self.send_flag.is_set():
                time.sleep(0.1)
            self.close_event.set()
            self.mega_data_thread.join()

            print("[Launcher] : Drill COMPLETE!")
            print("======================================================================")
            print(self.player,"'s results for ", self.drillType, " passing drill:")
            print("User selcted a Difficulty: ", self.difficulty, " and a Speed: ", self.drillSpeed)
            print()
            print("Ball #               :          one    |    two    |     three    |    four    |    five    ")
            try:
                print("Player reaction time :        ", self.targetTimes[1], "    ", self.targetTimes[2], "    ",
                      self.targetTimes[3], "    ", self.targetTimes[4], "    ", self.targetTimes[5], "    ")
                print("Player passing speed :        ", self.ballSpeeds[1], "    ", self.ballSpeeds[2], "    ",
                      self.ballSpeeds[3], "    ", self.ballSpeeds[4], "    ", self.ballSpeeds[5], "    ")
                print("======================================================================")


            except AttributeError as a:
                pass
            try:
                print("At a current temperature of :        ", self.drill_data.temp)
            except:
                pass

            self.output_to_excel()  # NEW LINEs *!*!*!*!*!*!**!*!*!*!*!*!*!
            self.display_results()

            self.kill_event.set()
            sys.exit()

        else:
            print("[Launcher] : Not sure what went wrong")
            self.send_data = '<-1,-1,0,6,0,0.0>'
            send_mega_stack.push(self.send_data)
            self.send_flag.set()
            while self.send_flag.is_set() or self.py_reset.is_set:
                time.sleep(0.1)
            sys.exit()

    def run(self):
        print("[Launcher] : Starting ...")
        self.launcher_startup()

        try:
            if self.drillType == 'Dynamic':
                # ____________________________ DYNAMIC DRILL _____________________________ #
                self.dynamic_drill()

            if self.drillType == 'Static':
                # ____________________________ STATIC DRILL _____________________________ #
                self.static_drill()

            if self.drillType == 'Manual':
                # ____________________________ MANUAL DRILL _____________________________ #
                self.manual_drill()

        except Exception as e:
            print('[Launcher] : failed because of exception: ' + str(e))
